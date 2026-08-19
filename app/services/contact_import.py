"""Turn a spreadsheet into rows the dial pump can work from.

Lead lists arrive as whatever the last agency sent: a header row that may or may not be the
first row, a phone column called "Mobile" or "Contact No." or "phone", numbers stored as text
with spaces or as floats Excel has helpfully rounded, blank rows in the middle, and the same
person twice.

Two decisions shape the whole module.

**Nothing is rejected wholesale.** An import that fails because row 400 of 3,000 has a bad
number is an import the operator cannot use — they have no way to find row 400. Every row is
either accepted or reported with its own line number and reason, and the good rows still land.

**Reading the file is separate from writing it.** parse() touches no database, so the same
code path produces the preview an operator confirms and the rows that are then written. A
preview computed by different logic than the commit is a preview that lies.

Read-only streaming mode throughout: a 20,000-row sheet is a normal lead list, and openpyxl's
default mode builds the entire workbook in memory before you can look at a single cell.
"""

import uuid
from dataclasses import dataclass, field
from typing import IO, Dict, List, Optional, Tuple

from loguru import logger
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.db import Contact, ContactStatus, Suppression
from app.utils.phone import to_e164

# Header spellings seen in real lead lists, lowercased and stripped of punctuation. Matched by
# containment, so "Customer Mobile No." finds "mobile".
#
# Ordered most specific first, and the best match wins rather than the leftmost. A sheet with
# both "Alt Number" and "Mobile" is ordinary, and taking the first hit picks the alternate
# number — so every call goes to the line the prospect gave as a fallback.
_PHONE_HEADERS = ("mobile", "phone", "whatsapp", "cell", "contact", "number", "tel")
_NAME_HEADERS = ("name", "customer", "client", "lead", "prospect", "person")

# A header carrying one of these is a secondary number even when it also says "mobile", so it
# loses to any column that does not.
_SECONDARY_MARKERS = ("alt", "alternate", "secondary", "other", "spouse", "office")


def _phone_rank(label: str) -> Optional[int]:
    """How good a phone-column header this is. Lower is better; None means not one."""
    for rank, token in enumerate(_PHONE_HEADERS):
        if token in label:
            # Pushed behind every primary match, but still ahead of no match at all: a sheet
            # whose only number column is "Alt Mobile" should import rather than fail.
            return rank + (len(_PHONE_HEADERS) if any(m in label for m in _SECONDARY_MARKERS) else 0)
    return None

# How far down to look for the header row. Exports commonly carry a title and a blank line
# above it; beyond a handful of rows it is not a header, it is data with no header at all.
_HEADER_SEARCH_ROWS = 10

# Rows accepted from one file. Above this the request itself is the problem — a browser upload
# that takes minutes will be retried by an impatient operator, and each retry re-parses.
MAX_IMPORT_ROWS = 20_000


@dataclass
class RowProblem:
    """A row that will not be dialled, and why, in terms of the file the operator has open."""

    row: int
    reason: str
    value: str = ""


@dataclass
class ParsedContact:
    row: int
    phone_number: str
    name: Optional[str] = None


@dataclass
class ImportReport:
    """What an import did, or would do. The same shape either way."""

    batch_id: uuid.UUID
    # Column letters the headers were found in, so the operator can see it read the file the
    # way they meant. A silent mis-mapping is the failure this exists to prevent.
    phone_column: Optional[str] = None
    name_column: Optional[str] = None
    header_row: Optional[int] = None
    total_rows: int = 0
    contacts: List[ParsedContact] = field(default_factory=list)
    problems: List[RowProblem] = field(default_factory=list)
    # Filled in by load(); meaningless on a dry run and left at zero there.
    inserted: int = 0
    already_present: int = 0
    suppressed: int = 0

    @property
    def dialable(self) -> int:
        return len(self.contacts)


def _normalise(header) -> str:
    return "".join(ch for ch in str(header or "").lower() if ch.isalnum() or ch == " ").strip()


def _looks_like_a_number(value) -> bool:
    """Enough digits to be a phone number rather than a serial or a pincode."""
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    return len(digits) >= 10


def find_columns(rows: List[tuple]) -> Tuple[Optional[int], Optional[int], Optional[int]]:
    """(header_row_index, phone_index, name_index) over the first few rows.

    Header names first, because they are what the operator meant. Failing that, the shape of
    the data: a column where most values have ten or more digits is the phone column whatever
    it is called, and that is what makes a headerless export work.
    """
    for index, row in enumerate(rows[:_HEADER_SEARCH_ROWS]):
        phone = name = None
        best = None
        for column, cell in enumerate(row):
            label = _normalise(cell)
            if not label:
                continue
            rank = _phone_rank(label)
            if rank is not None:
                if best is None or rank < best:
                    best, phone = rank, column
                continue
            if name is None and any(h in label for h in _NAME_HEADERS):
                name = column
        if phone is not None:
            return index, phone, name

    # No recognisable header. Find the column that holds phone numbers by looking at values.
    body = rows[:50]
    width = max((len(r) for r in body), default=0)
    for column in range(width):
        values = [r[column] for r in body if column < len(r) and r[column] not in (None, "")]
        if values and sum(_looks_like_a_number(v) for v in values) >= max(1, len(values) // 2):
            # A name column, if there is one, is the first other column carrying text.
            name = next(
                (
                    c
                    for c in range(width)
                    if c != column
                    and any(
                        isinstance(r[c], str) and r[c].strip()
                        for r in body
                        if c < len(r)
                    )
                ),
                None,
            )
            return None, column, name
    return None, None, None


def _cell_to_text(value) -> str:
    """A phone number as the sheet stored it, not as Python repr'd it.

    openpyxl already hands back an integral numeric cell as int, so str() is right for those.
    A float that is NOT integral is the case that matters: repr(9876543217.5) is
    "9876543217.5", to_e164 strips the dot as ordinary punctuation, and the result is
    +9198765432175 — a number one digit long and belonging to somebody else. is_fraction()
    below is what stops that reaching the dialler.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, (int, float)):
        return repr(value) if isinstance(value, float) else str(value)
    return str(value).strip()


def is_fraction(value) -> bool:
    """True for a numeric cell with a fractional part.

    Not a phone number, and dangerous rather than merely wrong: the fractional point is
    stripped as punctuation and the digits either side run together into a valid-looking
    number for a different person.
    """
    return isinstance(value, float) and not value.is_integer()


def parse(stream: IO[bytes]) -> ImportReport:
    """Read the sheet. Touches no database, so the preview and the commit agree by construction."""
    report = ImportReport(batch_id=uuid.uuid4())

    # data_only so formula cells yield their last computed value rather than "=A1&B1".
    workbook = load_workbook(stream, read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(row)
            if len(rows) > MAX_IMPORT_ROWS + _HEADER_SEARCH_ROWS:
                break
    finally:
        # read_only keeps file handles open until this is called.
        workbook.close()

    header_index, phone_column, name_column = find_columns(rows)
    if phone_column is None:
        report.problems.append(
            RowProblem(
                row=0,
                reason=(
                    "No phone column found. Give one column a header containing 'phone', "
                    "'mobile' or 'number', or make sure the numbers are in a column of "
                    "their own."
                ),
            )
        )
        return report

    from openpyxl.utils import get_column_letter

    report.header_row = None if header_index is None else header_index + 1
    report.phone_column = get_column_letter(phone_column + 1)
    report.name_column = None if name_column is None else get_column_letter(name_column + 1)

    body_from = 0 if header_index is None else header_index + 1
    seen: Dict[str, int] = {}

    for offset, row in enumerate(rows[body_from:]):
        line = body_from + offset + 1  # spreadsheet rows are 1-based
        if len(report.contacts) + len(report.problems) >= MAX_IMPORT_ROWS:
            report.problems.append(
                RowProblem(
                    row=line,
                    reason=f"Stopped at {MAX_IMPORT_ROWS} rows. Split the file and import again.",
                )
            )
            break

        cell = row[phone_column] if phone_column < len(row) else None
        raw = _cell_to_text(cell)
        name = ""
        if name_column is not None and name_column < len(row):
            name = str(row[name_column] or "").strip()

        # A wholly blank row is a spreadsheet artefact, not a problem to report.
        if not raw and not name:
            continue
        report.total_rows += 1

        if not raw:
            report.problems.append(RowProblem(row=line, reason="No phone number", value=name))
            continue
        if is_fraction(cell):
            report.problems.append(
                RowProblem(
                    row=line,
                    reason="This cell holds a decimal, not a phone number. Format the column "
                    "as Text and export again.",
                    value=raw,
                )
            )
            continue
        if "e+" in raw.lower():
            report.problems.append(
                RowProblem(
                    row=line,
                    reason="Excel stored this as scientific notation and the digits are lost. "
                    "Format the column as Text and export again.",
                    value=raw,
                )
            )
            continue

        try:
            number = to_e164(raw)
        except ValueError as e:
            report.problems.append(RowProblem(row=line, reason=str(e), value=raw))
            continue

        if number in seen:
            report.problems.append(
                RowProblem(
                    row=line,
                    reason=f"Same number as row {seen[number]} in this file",
                    value=number,
                )
            )
            continue
        seen[number] = line
        report.contacts.append(ParsedContact(row=line, phone_number=number, name=name or None))

    return report


async def load(db: AsyncSession, campaign_id, report: ImportReport) -> ImportReport:
    """Write a parsed report's contacts to one campaign's queue.

    Idempotent. ON CONFLICT DO NOTHING against the (campaign_id, phone_number) unique
    constraint is what makes re-uploading the same sheet a no-op instead of a second round of
    calls to everybody on it — and an operator who is unsure whether the first upload worked
    will upload again.

    Numbers already suppressed are written as DND rather than skipped. An operator looking at
    the campaign needs to see that the number was in their list and why it will not be
    dialled; a silently missing row looks like the import dropped it.
    """
    if not report.contacts:
        return report

    numbers = [c.phone_number for c in report.contacts]
    blocked = set(
        (
            await db.execute(
                select(Suppression.phone_number).where(Suppression.phone_number.in_(numbers))
            )
        )
        .scalars()
        .all()
    )

    rows = [
        {
            "id": uuid.uuid4(),
            "campaign_id": campaign_id,
            "phone_number": c.phone_number,
            "name": c.name,
            "status": ContactStatus.DND if c.phone_number in blocked else ContactStatus.PENDING,
            "last_outcome": "on the suppression list" if c.phone_number in blocked else None,
            "source_row": c.row,
            "import_batch_id": report.batch_id,
        }
        for c in report.contacts
    ]

    statement = (
        insert(Contact)
        .values(rows)
        .on_conflict_do_nothing(constraint="uq_contacts_campaign_phone")
        .returning(Contact.id)
    )
    inserted = len((await db.execute(statement)).scalars().all())
    await db.commit()

    report.inserted = inserted
    report.already_present = len(rows) - inserted
    report.suppressed = len(blocked)
    logger.info(
        f"Imported {inserted} contact(s) into campaign {campaign_id} "
        f"(batch {report.batch_id}); {report.already_present} already present, "
        f"{len(blocked)} suppressed, {len(report.problems)} row(s) rejected"
    )
    return report
